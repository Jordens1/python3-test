#！/usr/bin/env python
# _*_ coding:utf-8 _*_

import xlrd
#from xlwt import *
import requests, json
#from xlutils.copy import copy
from openpyxl import load_workbook
#import openpyxl
import time

#读取文件位置
def read_file(file_url):
    try:
        data = xlrd.open_workbook(file_url)
        return data
    except Exception as e:
        print(str(e))

#读取sheet       取出数值data   调用api取出结果    保存 分析数值  写入
def filter_excel(workbook, save_url, column_name=0, by_name='Sheet1'):
    table = workbook.sheet_by_name(by_name)  # 获得表格
    total_rows = table.nrows  # 拿到总共行数
    total_cols = table.ncols  # 拿到总列数
    excel_list = []
    for one_row in range(1, total_rows):  # 也就是从Excel第二行开始，第一行表头不算
        row_two = table.cell_value(one_row, 1)
        if row_two:
            init_url = "91d5dfef6c9f833adb66adccf75eb4a8"
            comp_url = "https://restapi.amap.com/v3/geocode/geo?key=%s&address=浙江省杭州市%s" % ( init_url, row_two)  #url拼接
            get_data = requests.get(comp_url).text         #获取返回值
            # excel_list.append(row_object)
            get_jsondata = json.loads(get_data)            #转为json格式
            jinweidu = get_jsondata["geocodes"][0]["location"]        #处理数据
            weidu = jinweidu.split(",")[0]
            jindu = jinweidu.split(",")[1]
            # w = copy(save_url)
            #             # w.get_sheet().write()
            #             # w.save(2 + save_url)
            wb = load_workbook(save_url)  # 生成一个已存在的wookbook对象
            wb1 = wb.active  # 激活sheet
            wb1.cell(one_row + 1, 4, jindu)  # 往sheet中的第二行第二列写入‘pass2’的数据
            wb1.cell(one_row + 1, 5, jindu)  # 往sheet中的第二行第二列写入‘pass2’的数据
            wb.save(save_url)  # 保存
            time.sleep(0.001)


    #return excel_list


file_url = r"D:\点位信息.xlsx"
workbook = read_file(file_url)
filter_excel(workbook, file_url)















