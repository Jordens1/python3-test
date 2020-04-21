#！/usr/bin/env python
# _*_ coding:utf-8 _*_

import xlrd
import requests, json
from openpyxl import load_workbook
import time

#读取文件位置
def read_file(file_url):
    try:
        data = xlrd.open_workbook(file_url)
        return data
    except Exception as e:
        print(str(e))

#读取sheet       取出数值data   调用api取出结果    保存 分析数值  写入
def filter_excel(workbook, by_name='Sheet1'):
    table = workbook.sheet_by_name(by_name)  # 获得表格
    total_rows = table.nrows  # 拿到总共行数
    #total_cols = table.ncols  # 拿到总列数
    weidu_total = []
    jingdu_total = []
    for one_row in range(1, total_rows):  # 也就是从Excel第二行开始，第一行表头不算
        row_two = table.cell_value(one_row, 1)
        if row_two:
            province = "浙江省杭州市"
            init_url = "91d5dfef6c9f833adb66adccf75eb4a8"
            comp_url = "https://restapi.amap.com/v3/geocode/geo?key=%s&address=%s%s" % (init_url,  province, row_two)  #url拼接
            get_data = requests.get(comp_url).text         #获取返回值
            get_jsondata = json.loads(get_data)            #转为json格式
            jinweidu = get_jsondata["geocodes"][0]["location"]        #处理数据
            weidu = jinweidu.split(",")[0]
            jingdu = jinweidu.split(",")[1]
            weidu_total.append(weidu)
            jingdu_total.append(jingdu)
    return total_rows, weidu_total, jingdu_total

def data_write(total_rows, weidu, jingdu, save_url):
        wb = load_workbook(save_url)  # 生成一个已存在的wookbook对象
        wb1 = wb.active  # 激活sheet
        for write_cow in range(2, total_rows+1):
            wb1.cell(write_cow, 4, jingdu[write_cow-2])  # 往sheet1中的第write_cow行第4列写入数据
            wb1.cell(write_cow, 5, weidu[write_cow-2])  # 往sheet1中的第write_cow行第5列写入数据
            wb.save(save_url)  # 保存
            time.sleep(0.001)

if __name__ == '__main__':
    file_url = r"D:\点位信息.xlsx"          #给的文件的路径
    workbook_data = read_file(file_url)
    total_rows, wedu_total, jindu_total= filter_excel(workbook_data)
    data_write(total_rows, wedu_total, jindu_total, file_url)