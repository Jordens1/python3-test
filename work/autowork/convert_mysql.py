#/usr/bin/env python
# _*_ coding:utf-8 _*_

import os, re
from openpyxl import Workbook
import sys

def conver(path, dest_filename):
    with open(path, 'r', encoding='utf-8') as f:
        id_pa = re.compile( r"device_id: ([0-9]+)", re.S )
        name_pa = re.compile( r"name: (\w+)", re.S )
        id = id_pa.findall(f.read())
        f.seek(0)
        name = name_pa.findall(f.read())
        #创建一个新的execl，需要save
        wb = Workbook()
        ws = wb.active
        #改sheet名字
        ws.title = "点位"
        #插入数据
        ws.cell(row=1, column=1, value='device_id')
        ws.cell(row=1, column=2, value='device_name')
        for i in range(2, len(id)+1):
            ws.cell(row=i, column=1, value=id[i-1])
            ws.cell(row=i, column=2, value=name[i-1])
        #保存为
        wb.save( dest_filename )

if __name__ == '__main__':
    txtpath = os.getcwd() + '\\' + sys.argv[1]
    dest_filename = '点位.xlsx'
    conver(txtpath, dest_filename)
