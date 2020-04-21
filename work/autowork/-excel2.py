#/usr/bin/env python
# _*_ coding:utf-8 _*_
from openpyxl.reader.excel import load_workbook

#把所有表的信息拿出来
def readExcel(path):
    file = load_workbook(filename=path)
    sheetName = file.sheetnames
    dic = {}
    for evesheet in sheetName:
        sheetInfo = []
        sheet = file[evesheet]
        rowConut = sheet.max_row
        columnConut = sheet.max_column
        for lineNum in range(1, rowConut + 1):
            linelist = []
            for couNum in range(1, columnConut + 1):
                value = sheet.cell(row= lineNum, column= couNum).value
                if value != None:
                    linelist.append(value)
            sheetInfo.append(linelist)
        dic[evesheet] = sheetInfo
    return dic
path = r'C:\python3test\视在\autowork\test.xlsx'
info = readExcel(path)





