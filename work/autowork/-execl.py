#/usr/bin/env python
# _*_ coding:utf-8 _*_

'''
xlsx xls 两种格式的文件


'''
from openpyxl.reader.excel import load_workbook

def readExcel(path):
    #打开文件
    file = load_workbook(filename=path)
    #获取表格名称  第一个表格
    # sheetName = file.get_sheet_names()
    sheetName = file.sheetnames[0]
    #拿出一个表格   前面的已经弃用了
    # sheet = file.get_sheet_by_name(sheetName)
    sheet = file[sheetName]
    #有多少行
    rowConut = sheet.max_row
    #有多少列
    columnConut = sheet.max_column
    #表格名字
    sheetTitle = sheet.title

    #读取数据
    for lineNum in range(1, rowConut + 1):
        linelist = []
        for couNum in range(1, columnConut + 1):
            #拿数据
            value = sheet.cell(row= lineNum, column= couNum).value
            if value != None:
                linelist.append(value)
        print(linelist)

#这里是只能处理xlsx文件
path = r'C:\python3test\视在\autowork\test.xlsx'
readExcel(path)












